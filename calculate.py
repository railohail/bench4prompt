import json
from collections import defaultdict
from openpyxl import Workbook

def load_questions():
    with open('questions.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def process_data(data, questions):
    user_scores = defaultdict(lambda: defaultdict(list))
    
    for question_id, question_data in data.items():
        for user_data in question_data.values():
            for attempt in user_data:
                username = attempt['username']
                score = attempt['score']
                user_scores[username][question_id].append(score)
    
    result = {}
    all_question_ids = set(questions.keys())
    
    for username, questions_attempted in user_scores.items():
        total_score = 0
        num_questions_attempted = 0
        missed_questions = all_question_ids - set(questions_attempted.keys())
        
        for question_id, scores in questions_attempted.items():
            avg_score = sum(scores) / len(scores)
            total_score += avg_score
            num_questions_attempted += 1
        
        average_total = total_score / len(all_question_ids)
        
        result[username] = {
            "average_total": average_total,
            "questions_attempted": num_questions_attempted,
            "total_questions": len(all_question_ids)
        }
        
        if missed_questions:
            result[username]["missed_questions"] = list(missed_questions)
    
    return result

def create_excel_report(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "平均分數"
    
    ws['A1'] = "用戶名"
    ws['B1'] = "平均分數"
    
    row = 2
    for username, user_data in data.items():
        ws[f'A{row}'] = username
        ws[f'B{row}'] = user_data['average_total']
        row += 1
    
    wb.save("平均分數.xlsx")
    print("Excel report '平均分數.xlsx' has been created.")

def main():
    # Load the questions data
    questions = load_questions()
    
    # Load the user score data
    with open('leaderboard.json', 'r') as f:
        data = json.load(f)
    
    # Process the data
    result = process_data(data, questions)
    
    # Save the result to a JSON file
    with open('average_scores.json', 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    
    print(f"Results have been saved to 'average_scores.json'")
    
    # Create Excel report
    create_excel_report(result)

if __name__ == "__main__":
    main()